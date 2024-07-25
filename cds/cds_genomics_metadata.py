import yaml
import pandas as pd
import requests
import json
import os

import synapseclient
from google.cloud import bigquery
from openpyxl import load_workbook

def load_bq(client, project, dataset, table, data):
    '''
    Load dataframe to BigQuery
    '''
    print('Loading: '+dataset+'.'+table)
    table_bq = '%s.%s.%s' % (project, dataset, table)
    job_config = bigquery.LoadJobConfig( 
        write_disposition="WRITE_TRUNCATE",
        autodetect=False,
        source_format=bigquery.SourceFormat.CSV
    )
    
    job = client.load_table_from_dataframe(
        data, table_bq, job_config=job_config
    )


def get_cds_value_set(cde_pvs, cds_attribute):
    '''
    Function to map HTAN attributes to similar CDS attributes containing commas
    '''

    pv_map = pd.DataFrame()
    pv_map['CDS'] = cde_pvs
    pv_map['HTAN'] = [x.replace(',','').replace(';','') for x in cde_pvs]
    
    return pv_map


def main():

    template_type = 'CDS Genomics'
    package_id = 'v24.4.1.seq' #DCC-defined package ID (v<YY>.<M>.<X>.<seq|img>)
    phs_accession = 'phs002371.v6.p1' #dbGaP study version associated with this submission
    data_release = 'Release 5.0' #HTAN data release associated with this submission

    client = bigquery.Client()

    syn = synapseclient.Synapse()
    syn.login()

    print( ' ' )
    print( ' Preparing CDS template for %s ' % package_id )
    print( ' ' )

    # Read in CDS Excel Metadata Template
    url = "https://github.com/CBIIT/cds-model/raw/main/metadata-manifest/CDS_Metadata_Submission_Template_v1.3.3.xlsx"

    res = requests.get(url, allow_redirects=True)
    with open('../templates_and_dictionaries/%s' % url.split('/')[-1],'wb') as file:
        file.write(res.content)

    # Open workbook
    wb = load_workbook(filename = '../templates_and_dictionaries/%s' 
        % url.split('/')[-1])
    value_set = wb['Terms and Value Sets']
    
    pv_df = pd.DataFrame(value_set.values)
    pv_df = pv_df.rename(columns=pv_df.iloc[0]).loc[1:]

    # read in mappings yaml file
    with open('../mapping/mappings.yaml', 'r') as f:
        mappings = yaml.safe_load(f)

    # read in file containing center-specific info
    htan_center_mappings = json.load(
        open('../mapping/htan_center_mappings.json'))

    with open('./config.yaml', 'r') as file:
        config = yaml.safe_load(file)
   
    components = config['seq_components']

    files = pd.DataFrame()

    # exclude 3 Hi-C level 2 files that should be in level 3
    for c in components:
        df = client.query("""
            SELECT * FROM `htan-dcc.combined_assays.{}`
            WHERE Data_Release IS NOT NULL
            AND CDS_Release IS NULL
            AND File_Format NOT LIKE 'bai'
            AND Filename NOT LIKE '%.bai'
            AND entityId NOT IN 
            ('syn39637703','syn39637986','syn39638039')
        """.format(c)).result().to_dataframe()
        files = pd.concat([files,df])

    id_prov = client.query("""
        SELECT DISTINCT entityId, HTAN_Participant_ID, 
        HTAN_Assayed_Biospecimen_ID AS HTAN_Biospecimen_ID 
        FROM `htan-dcc.id_provenance.upstream_ids`
    """).result().to_dataframe()

    # get biospecimen and clinical metadata
    clinical = client.query("""
        SELECT DISTINCT * EXCEPT(HTAN_Participant_ID)
        FROM `htan-dcc.metadata.cds_clinical`
        """).result().to_dataframe()

    # merge assay and clinical metadata
    full_metadata = files.merge(id_prov, how = 'left',
        on='entityId').merge(clinical, 
        how = 'left', on = 'HTAN_Biospecimen_ID'
    )

    # add PI names and emails
    for n in ['pi_first', 'pi_last', 'pi_email', 'bucket']:
        full_metadata[n] = [
        htan_center_mappings[y][n] for y in [
        x.split('_')[0] for x in list(
            full_metadata['HTAN_Data_File_ID'])]
    ]

    full_metadata['transfer_id'] = package_id
    
    full_metadata['file_url_in_cds'] = [
        's3://cds-243-phs002371/'+x+'/'+y+'/'+z.replace(' ','') for 
        x,y,z, in zip(list(full_metadata['transfer_id']),
        [x.split('/')[2:3][0] for x in list(
            full_metadata['bucket'])],
        [x.replace(' ','') for x in list(full_metadata['Filename'])])
    ]

    # convert age at diagnosis to years
    full_metadata['age_at_diagnosis_years'] = [int(float(x)/365) 
        if x not in [None, 'unknown', 'Not Applicable', 'Not Reported'] and
        pd.isna(x) == False
        else None for x in list(full_metadata['Age_at_Diagnosis'])]

    # mask ages <18 and >89
    full_metadata['age_at_diagnosis_years'] = [90 if y>89 else 18 
        if y<18 else y for y in list(
        full_metadata['age_at_diagnosis_years'])]

    # table - useful for internal tracking 
    full_metadata.to_csv(
        '../tables/%s/full_metadata_%s.csv'
        % (package_id, package_id),
        index=False
    )

    num_rows = len(full_metadata)
    cds = pd.DataFrame([phs_accession] * num_rows,
        columns = ['phs_accession']
    )

    for a in mappings[template_type]['attributes']:
        target_attribute = a['target_attribute']
        source_attribute = a['source_attribute']
    
        print(target_attribute)

        if target_attribute == 'phs_accession':
            continue
        elif 'fixed_value' in a:
            cds = cds.assign(**{
                target_attribute: [a['fixed_value']] * num_rows
            })
        elif 'dict' in a:
            r = eval(a['map'], {
                'source_col': full_metadata[
                    source_attribute.replace(' ','_')],
                'dict': dict(a['dict'])
                }
            )
            cds = cds.assign(**{target_attribute: r})
        elif 'map' in a:
            r = eval(a['map'], {
                'source_col': full_metadata[
                    source_attribute.replace(' ','_')],
                }
            )
            
            if target_attribute in [
                'primary_diagnosis','site_of_resection_or_biopsy']:
                    i_s = int(pv_df.index[pv_df['Value Set Name'] == target_attribute].tolist()[0])
                    i_e = pv_df['Value Set Name'].iloc[i_s:].dropna().index[0]
                    template_pv_list = [x for x in list(pv_df['Term'][i_s-1:i_e-1]) if x is not None]

                    pv_map = get_cds_value_set(template_pv_list,target_attribute)
                    r = [pv_map[pv_map['HTAN']==x]['CDS'].values[0] if x in list(
                        pv_map['HTAN']) else x for x in r]
            
            cds = cds.assign(**{target_attribute: r})
        
        else:
            cds = cds.assign(**{target_attribute: None})

    cds = cds.loc[cds.astype(str).drop_duplicates().index]

    # save as excel workbook per CDS' request
    if 'Metadata' in wb.sheetnames:
    # if 'Metadata' sheet exists in workbook, remove and replace with new df
        metadata_sheet_index = wb.sheetnames.index('Metadata')
        wb.remove(wb.worksheets[metadata_sheet_index])
       
        metadata_sheet = wb.create_sheet('Metadata', 
            index=metadata_sheet_index)
        header_row = list(cds.columns) 
        metadata_sheet.append(header_row)
        
        for row in cds.itertuples(index=False):
            metadata_sheet.append(list(row))
    
    else:
        metadata_sheet = wb.create_sheet('Metadata')
        metadata_sheet.append(cds.columns.tolist())
        
        for row in cds.itertuples(index=False):
            metadata_sheet.append(list(row))

    wb.save('../tables/%s/%s_%s.xlsx' 
        % (package_id, template_type.replace(" ", "_"), package_id))

    
    # -------------------------------------------------------------------
    # Create updated Subject Sample Mapping table for dbGaP submission

    print ( '' )
    print ( 'Preparing Subject Sample Mapping and Subject Consent tables for dbGaP' )
    print ( '' )

    ssm = client.query("""
        SELECT * FROM `htan-dcc.cds.dbGaP_SSM`
    """).result().to_dataframe()

    bios = client.query("""
        SELECT DISTINCT HTAN_Biospecimen_ID, 
        Id AS CDS_SAMPLE_ID
        FROM `htan-dcc.combined_assays.Biospecimen`
        """).result().to_dataframe()

    HTAN_SSM_DS = full_metadata[[
        'HTAN_Participant_ID','HTAN_Biospecimen_ID']
    ].drop_duplicates()

    new_ssm = HTAN_SSM_DS.merge(bios, how='left', 
        on='HTAN_Biospecimen_ID').rename(columns={
            "HTAN_Participant_ID" : "SUBJECT_ID", 
            "HTAN_Biospecimen_ID" : "SAMPLE_ID"}
        )

    full_ssm = pd.concat(
        [ssm,new_ssm],ignore_index=True).drop_duplicates(
            subset = ['SAMPLE_ID'], keep = 'first'
        )

    # output csv file to submit to dbgap portal
    full_ssm.to_csv(
        '../tables/%s/HTAN_SSM_DS_%s.csv' 
        % (package_id,package_id), index=False
    )
    
    # update bigquery SSM tables

    load_bq(
        client, 'htan-dcc', 'cds', 'dbGaP_SSM', full_ssm)

    load_bq(
        client, 'htan-dcc', 'cds', 
        'dbGaP_SSM_%s' % phs_accession.replace('.','_'), full_ssm)

    # ----------------------------------------------------------------
    # Create new Subject Consent table for dbGaP submission

    sc = client.query("""
        SELECT * FROM `htan-dcc.cds.dbGaP_SC`
    """).result().to_dataframe()

    HTAN_SubjectConsent_DS = full_metadata[
        ['HTAN_Participant_ID','Gender']].rename(
            columns={
                'HTAN_Participant_ID': 'SUBJECT_ID',
                'Gender': 'SEX'
            }
        ).drop_duplicates()

    HTAN_SubjectConsent_DS["CONSENT"] = [
        htan_center_mappings[x.split("_")[0]]['consent'] 
        for x in list(HTAN_SubjectConsent_DS['SUBJECT_ID'])
    ]

    add = pd.concat([HTAN_SubjectConsent_DS,sc])
    add["SEX"] = add['SEX'].str.replace(
        'Not Reported', 'Unknown', regex=True
    )
    add["SEX"] = [x.capitalize() if x is not None else 'Unknown' 
        for x in list(add["SEX"])]

    # set column order required for dbGaP pipelines
    add = add[['SUBJECT_ID','CONSENT','SEX']].drop_duplicates()

    # output csv file to submit to dbgap portal
    add.to_csv(
        '../tables/%s/HTAN_SubjectConsent_DS_%s.csv'
        % (package_id,package_id),
        index=False)

    # update bigquery SC table
    load_bq(
        client, 'htan-dcc', 'cds', 'dbGaP_SC', add)

    load_bq(
        client, 'htan-dcc', 'cds', 
        'dbGaP_SC_%s' % phs_accession.replace('.','_'), add)

    # ---------------------------------------------------------------------------
    # Output text to include in CDS change log

    overlap = len(set(sc['SUBJECT_ID']) &
        set(HTAN_SubjectConsent_DS["SUBJECT_ID"]))

    # get participant count for cds release notes
    centers = ', '.join([x.replace('HTAN ','') for x in list(
        set(full_metadata['HTAN_Center']))])

    print ( '' )
    print ( 'Done!' )
    print ( '' )

    print( 'Change log narrative: ')
    print ( '' )

    print ( 'New Updates: %s files from %s patient cases (%s new) were added to the bucket' 
        % (len(set(cds['file_url_in_cds'])),
            len(HTAN_SubjectConsent_DS.drop_duplicates()),
            str(len(HTAN_SubjectConsent_DS.drop_duplicates())-overlap)))
    
    print ( 'Data was submitted by %s HTAN centers' 
        % centers)
    print ( '' )
    print ( 'Narrative: This transfer contains genomics data submitted by %s HTAN centers. %s level 1 & 2 sequencing files were added from %s patient cases (%s new) participants' 
        % (centers,
        len(set(cds['file_url_in_cds'])),
        len(HTAN_SubjectConsent_DS.drop_duplicates()),
        str(len(HTAN_SubjectConsent_DS.drop_duplicates())-overlap)))
    print ( '' )
 

    # Also need to update CDS_Release column of 
    # htan-dcc.released.entities table after files are released on CDS

if __name__ == "__main__":
    main()
